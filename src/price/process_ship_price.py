import logging
import sys
from pathlib import Path
from typing import Dict, Any

import openpyxl

from src.summary.process_summary import load_summary_shipments, summarize_shipment


logger = logging.getLogger(__name__)

SHIP_PRICE_PATH = Path("./shipment_data/ship_price.xlsx")
PRICE_MODES = ("FIST", "SEND")


def clean_value(value):
    return "" if value is None else str(value).strip()


def parse_number(value, field_name):
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        sys.exit(f"ship_price.xlsx 中 {field_name} 不是有效数字: {value!r}")


def money(value):
    return round(float(value), 2)


def format_money(value):
    return f"{money(value):.2f}"


def load_price_sheet(ws):
    carriers = []
    for col in range(2, ws.max_column + 1):
        carrier = clean_value(ws.cell(1, col).value)
        if not carrier:
            continue
        customs_fee = parse_number(ws.cell(2, col).value, f"{ws.title}-{carrier}-报关费")
        if customs_fee is None:
            sys.exit(f"ship_price.xlsx 中 {ws.title}-{carrier} 报关费为空。")
        carriers.append({"name": carrier, "col": col, "customs_fee": customs_fee})

    if not carriers:
        sys.exit(f"ship_price.xlsx 的 {ws.title} sheet 未找到承运商。")

    warehouses = []
    prices = {}
    for row in range(3, ws.max_row + 1):
        warehouse = clean_value(ws.cell(row, 1).value)
        if not warehouse:
            continue
        warehouses.append(warehouse)
        prices[warehouse] = {}
        for carrier in carriers:
            prices[warehouse][carrier["name"]] = parse_number(
                ws.cell(row, carrier["col"]).value,
                f"{ws.title}-{warehouse}-{carrier['name']}单价",
            )

    return {
        "carriers": carriers,
        "warehouses": warehouses,
        "prices": prices,
    }


def load_ship_price(path: Path = SHIP_PRICE_PATH):
    if not path.exists():
        sys.exit(f"未找到海运费价格表：{path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    missing_sheets = [mode for mode in PRICE_MODES if mode not in wb.sheetnames]
    if missing_sheets:
        sys.exit(f"ship_price.xlsx 缺少 sheet：{', '.join(missing_sheets)}")

    return {mode: load_price_sheet(wb[mode]) for mode in PRICE_MODES}


def build_mode_summaries(mode: str, product_dict: Dict[str, Dict], package_dict: Dict[str, Dict]):
    shipments = load_summary_shipments(mode, product_dict)
    summaries = {}
    for summary_key, ship_data in shipments.items():
        warehouse = summary_key.split("-", 1)[1]
        summaries[warehouse] = summarize_shipment(ship_data, package_dict)
    return summaries


def calculate_fee(summary: Dict[str, Any], unit_price: float, customs_fee: float):
    return money(float(summary.get("总重量(kg)", 0) or 0) * unit_price + customs_fee)


def calc_fist(fist_summary: Dict[str, Dict[str, Any]], fist_price: Dict[str, Any]):
    logger.info("▶ FIST 海运费测算")

    total_fee = 0.0
    output_warehouses = set()
    selections = []

    for warehouse in fist_price["warehouses"]:
        summary = fist_summary.get(warehouse)
        if not summary:
            continue

        candidates = []
        for carrier in fist_price["carriers"]:
            unit_price = fist_price["prices"].get(warehouse, {}).get(carrier["name"])
            if unit_price is None:
                continue
            fee = calculate_fee(summary, unit_price, carrier["customs_fee"])
            candidates.append({
                "carrier": carrier["name"],
                "customs_fee": carrier["customs_fee"],
                "unit_price": unit_price,
                "fee": fee,
            })

        output_warehouses.add(warehouse)
        if not candidates:
            logger.warning(f"FIST-{warehouse} 没有任何承运商报价，跳过该仓库的计算")
            continue

        best_fee = min(item["fee"] for item in candidates)
        best_candidates = [item for item in candidates if item["fee"] == best_fee]
        total_fee += best_fee
        best_details = []
        for best in best_candidates:
            detail = {
                "warehouse": warehouse,
                "summary": summary,
                "carrier": best["carrier"],
                "customs_fee": best["customs_fee"],
                "unit_price": best["unit_price"],
                "fee": best["fee"],
            }
            selections.append(detail)
            best_details.append(detail)
        log_fist_best_details(warehouse, best_details)

    for warehouse in sorted(set(fist_summary) - output_warehouses):
        logger.warning(f"FIST-{warehouse} 没有任何承运商报价，跳过该仓库的计算")

    total_fee = money(total_fee)
    logger.info(f"FIST 预计总费用: {format_money(total_fee)}元")
    return {
        "total_fee": total_fee,
        "selections": selections,
    }


def log_fist_detail(detail):
    summary = detail["summary"]
    logger.info(
        f"【FIST-{detail['warehouse']}】承运商: {detail['carrier']}，"
        f"总箱数： {summary.get('总箱数')} 箱，"
        f"重量: {summary.get('总重量(kg)')} kg，"
        f"单价: {detail['unit_price']} 元/kg，"
        f"报关费: {detail['customs_fee']}元，"
        f"费用: {format_money(detail['fee'])}元"
    )


def log_fist_detail_indented(detail):
    summary = detail["summary"]
    logger.info(
        f"    承运商: {detail['carrier']}，"
        f"总箱数：{summary.get('总箱数')} 箱，"
        f"重量: {summary.get('总重量(kg)')} kg，"
        f"单价: {detail['unit_price']} 元/kg，"
        f"报关费: {detail['customs_fee']}元，"
        f"费用: {format_money(detail['fee'])}元"
    )


def log_fist_best_details(warehouse, details):
    if len(details) == 1:
        log_fist_detail(details[0])
        return

    logger.info(f"【FIST-{warehouse}】存在多个最低费用承运商:")
    for detail in details:
        log_fist_detail_indented(detail)


def calc_send(send_summary: Dict[str, Dict[str, Any]], send_price: Dict[str, Any]):
    logger.info("▶ SEND 海运费测算")

    results = []
    for carrier in send_price["carriers"]:
        missing_warehouses = [
            warehouse for warehouse in send_summary
            if send_price["prices"].get(warehouse, {}).get(carrier["name"]) is None
        ]
        if missing_warehouses:
            for warehouse in missing_warehouses:
                logger.warning(f"{carrier['name']}承运商{warehouse}仓库报价缺失")
            continue

        details = []
        total_fee = 0.0
        for warehouse, summary in send_summary.items():
            unit_price = send_price["prices"][warehouse][carrier["name"]]
            fee = calculate_fee(summary, unit_price, carrier["customs_fee"])
            total_fee += fee
            details.append({
                "warehouse": warehouse,
                "summary": summary,
                "unit_price": unit_price,
                "customs_fee": carrier["customs_fee"],
                "fee": fee,
            })

        results.append({
            "carrier": carrier["name"],
            "total_fee": money(total_fee),
            "details": details,
        })

    if not send_summary:
        logger.info("暂无 SEND 货件数据")
        return []

    if not results:
        logger.info("暂无可计算的 SEND 承运商报价")
        return []

    results = sorted(results, key=lambda item: item["total_fee"])
    for result in results:
        logger.info(f"【{result['carrier']}】预计总费用: {format_money(result['total_fee'])}元")
        for detail in result["details"]:
            summary = detail["summary"]
            logger.info(
                f"    【{detail['warehouse']}】总箱数： {summary.get('总箱数')} 箱，"
                f"重量: {summary.get('总重量(kg)')} kg，"
                f"单价: {detail['unit_price']} 元/kg，"
                f"报关费: {detail['customs_fee']}元，"
                f"费用: {format_money(detail['fee'])}元"
            )
    return results


def log_final_recommendation(fist_result, send_results):
    logger.info("-------------------------------------")
    logger.info("▶ 最终建议方案")

    has_fist = fist_result and fist_result.get("selections")
    has_send = bool(send_results)
    if not has_fist and not has_send:
        logger.info("暂无可推荐方案")
        return

    if has_fist and not has_send:
        logger.info("推荐使用：FIST 分仓最优承运商组合")
        logger.info(f"预计总费用: {format_money(fist_result['total_fee'])}元")
        logger.info("-------------------------------------")
        for detail in fist_result["selections"]:
            log_fist_detail(detail)
        return

    best_send = send_results[0] if has_send else None
    if has_send and not has_fist:
        logger.info(f"推荐使用：SEND - {best_send['carrier']}")
        logger.info(f"预计总费用: {format_money(best_send['total_fee'])}元")
        return

    fist_fee = fist_result["total_fee"]
    send_fee = best_send["total_fee"]
    if send_fee < fist_fee:
        logger.info(f"推荐使用：SEND - {best_send['carrier']}")
        logger.info(f"预计总费用: {format_money(send_fee)}元")
        logger.info(f"对比 FIST 预计总费用: {format_money(fist_fee)}元")
        logger.info(f"预计节省: {format_money(fist_fee - send_fee)}元")
    elif fist_fee < send_fee:
        logger.info("推荐使用：FIST 分仓最优承运商组合")
        logger.info(f"预计总费用: {format_money(fist_fee)}元")
        logger.info(f"对比 SEND 最优方案【{best_send['carrier']}】: {format_money(send_fee)}元")
        logger.info(f"预计节省: {format_money(send_fee - fist_fee)}元")
        logger.info("-------------------------------------")
        for detail in fist_result["selections"]:
            log_fist_detail(detail)
    else:
        logger.info(f"FIST 分仓最优承运商组合 与 SEND - {best_send['carrier']} 预计总费用相同")
        logger.info(f"预计总费用: {format_money(fist_fee)}元")


def run_calc(product_dict: Dict[str, Dict], package_dict: Dict[str, Dict]):
    prices = load_ship_price()
    fist_summary = build_mode_summaries("FIST", product_dict, package_dict)
    send_summary = build_mode_summaries("SEND", product_dict, package_dict)

    fist_result = calc_fist(fist_summary, prices["FIST"])
    send_results = calc_send(send_summary, prices["SEND"])
    log_final_recommendation(fist_result, send_results)
