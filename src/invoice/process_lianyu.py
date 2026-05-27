import os
from datetime import datetime

import openpyxl

from src.constants.carriers import Carrier
from src.constants.dir_types import DirType
from src.utils.common_utils import root_dir
from src.utils.excel_utils import copy_row_style


def invoice_value(row, key):
    return row.get(key)


def split_cn_en(value):
    parts = str(value or "").split("/", 1)
    cn_value = parts[0].strip()
    en_value = parts[1].strip() if len(parts) > 1 else ""
    return cn_value, en_value


def number_value(value, field_name):
    try:
        return float(value)
    except (TypeError, ValueError):
        raise ValueError(f"联宇发票生成失败，{field_name} 不是有效数字: {value!r}")


def fill_lianyu_sheet(ws, global_info, package_dict, invoice_data, ship_data, start_row=2):
    warehouse_code = invoice_value(invoice_data, "仓库代码")
    material_cn, material_en = split_cn_en(global_info.get("Material"))
    current_row = start_row
    total_box_num = 0

    for product_type, data in (ship_data.get("list_data") or {}).items():
        if current_row != start_row:
            copy_row_style(ws, start_row, current_row)

        package_data = package_dict.get(product_type)
        if not package_data:
            raise ValueError(f"联宇发票生成失败，未找到套装类型 {product_type!r} 的包装信息")

        boxes = int(data.get("total_boxes") or 0)
        single_count = int(number_value(package_data.get("单箱数量"), "单箱数量"))
        declare_price = number_value(package_data.get("申报单价"), "申报单价")
        gross_weight = number_value(package_data.get("毛重(kg)"), "毛重(kg)")
        net_weight = number_value(package_data.get("净重(kg)"), "净重(kg)")

        total_box_num += boxes
        ws.cell(row=current_row, column=1, value=invoice_value(invoice_data, "货件ID"))
        ws.cell(row=current_row, column=3, value=invoice_value(invoice_data, "追踪编号"))
        # ws.cell(row=current_row, column=5, value=global_info.get("Model"))
        ws.cell(row=current_row, column=6, value=global_info.get("Product_Name_En"))
        ws.cell(row=current_row, column=7, value=global_info.get("Product_Name_Cn"))
        ws.cell(row=current_row, column=8, value=boxes)
        ws.cell(row=current_row, column=9, value=declare_price)
        ws.cell(row=current_row, column=10, value=single_count * boxes)
        ws.cell(row=current_row, column=11, value=material_en)
        ws.cell(row=current_row, column=12, value=material_cn)
        ws.cell(row=current_row, column=13, value=global_info.get("Usr_For"))
        ws.cell(row=current_row, column=14, value=global_info.get("HS_Code"))
        ws.cell(row=current_row, column=16, value=global_info.get("Brand"))
        ws.cell(row=current_row, column=19, value=round(gross_weight * boxes, 2))
        ws.cell(row=current_row, column=20, value=round(net_weight * boxes, 2))

        current_row += 1

    return warehouse_code, total_box_num


def fill_lianyu_template(global_info, package_dict, invoice_data, ship_data, start_row=2):
    date_str = datetime.today().strftime("%Y%m%d")
    template_path = f"./basic_data/template/{DirType.Invoice}/{Carrier.LianYu}.xlsx"
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]

    warehouse_code, total_box_num = fill_lianyu_sheet(ws, global_info, package_dict, invoice_data, ship_data, start_row)
    ws.title = f"{warehouse_code}-{total_box_num}"

    output_path = root_dir(DirType.Invoice) + f"/{Carrier.LianYu}_{warehouse_code}_{total_box_num}_{date_str}.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path
