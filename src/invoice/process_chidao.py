import os
import logging
from datetime import datetime
import openpyxl

from src.utils.common_utils import root_dir
from src.utils.excel_utils import insert_image_into_cell, clean_sheet_data, img_attr

logger = logging.getLogger(__name__)


def fill_chidao_sheet(ws, global_info, package_dict, invoice_data, ship_data, start_row=12):
    """
    ✅ 辅函数：对传入的 ws 执行“原单元格填充逻辑”（保持不变）
    返回 (repo_name, total_box_num)，用于 sheet 命名。
    """
    shipment_info = ship_data.get("shipment_info") or {}
    repo_name = shipment_info.get("仓库名称")

    ws["C4"] = "AMAZON-" + repo_name
    ws["C6"] = invoice_data[4]
    ws["C7"] = "有单证" if invoice_data[2] == "是" else "无单证"
    ws["C8"] = invoice_data[5]
    ws["C9"] = invoice_data[6]
    ws["C10"] = "深圳仓"

    image_errors = []
    ws.column_dimensions["C"].width = img_attr["col_width"]

    current_row = start_row
    total_box_num = 0

    for product_type, data in (ship_data.get("list_data") or {}).items():
        ws.cell(row=current_row, column=1, value=invoice_data[0])
        ws.cell(row=current_row, column=2, value=invoice_data[1])

        package_data = package_dict.get(product_type)
        insert_image_into_cell(ws, package_data, current_row, img_attr, image_errors, cell_anchor="C")

        ws.cell(row=current_row, column=4, value=global_info.get("HS_Code"))
        ws.cell(row=current_row, column=5, value=global_info.get("Product_Name_Cn"))
        ws.cell(row=current_row, column=6, value=global_info.get("Product_Name_En"))

        ws.cell(row=current_row, column=7, value=package_data.get("长(cm)"))
        ws.cell(row=current_row, column=8, value=package_data.get("宽(cm)"))
        ws.cell(row=current_row, column=9, value=package_data.get("高(cm)"))

        ws.cell(row=current_row, column=10, value=package_data.get("单箱数量"))

        boxes = data.get("total_boxes")
        total_box_num += boxes
        ws.cell(row=current_row, column=11, value=boxes)

        ws.cell(row=current_row, column=12, value=f"=J{current_row}*K{current_row}")  # L
        weight = package_data.get("毛重(kg)")
        ws.cell(row=current_row, column=13, value=round(float(weight) * int(boxes), 1))

        ws.cell(row=current_row, column=14, value=f"=G{current_row}*H{current_row}*I{current_row}*K{current_row}/1000000")  # N
        ws.cell(row=current_row, column=15, value=package_data.get("申报单价"))
        ws.cell(row=current_row, column=16, value=f"=L{current_row}*O{current_row}")  # P

        ws.cell(row=current_row, column=17, value=global_info.get("Material"))
        ws.cell(row=current_row, column=18, value=global_info.get("Usr_For"))
        ws.cell(row=current_row, column=20, value=global_info.get("Brand"))

        ws.cell(row=current_row, column=21, value=invoice_data[7])
        ws.cell(row=current_row, column=22, value="中国")

        current_row += 1

    if image_errors:
        logger.info("以下行图片加载失败：")
        for err in image_errors:
            logger.info(err)

    return repo_name, total_box_num

def fill_chidao_template(global_info, package_dict, invoice_data, ship_data, start_row=12):
    from src.constants.carriers import Carrier
    from src.constants.dir_types import DirType
    from src.invoice.process_ship_invoice import invoice_input_path

    date_str = datetime.today().strftime("%Y%m%d")
    output_path = root_dir(DirType.Invoice) + f"/{Carrier.ChiDao}_{date_str}.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # 1️⃣ 首次生成
    if not os.path.exists(output_path):
        wb = openpyxl.load_workbook(invoice_input_path(Carrier.ChiDao))
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]

        repo_name, total_box_num = fill_chidao_sheet(ws, global_info, package_dict, invoice_data, ship_data, start_row)
        ws.title = f"{repo_name}-{total_box_num}"
        sheets = wb._sheets

        # 新 sheet 放最前
        sheets.remove(ws)
        sheets.insert(0, ws)
        wb.save(output_path)
        logger.debug(f" 数据已保存到 {output_path}")
        return

    # 2️⃣ 追加 sheet
    wb = openpyxl.load_workbook(output_path)
    base_ws = wb.worksheets[0]
    new_ws = wb.copy_worksheet(base_ws)

    clean_sheet_data(new_ws, start_row=12, end_row=300, exclude_col=["L", "N", "P"])
    repo_name, total_box_num = fill_chidao_sheet(new_ws, global_info, package_dict, invoice_data, ship_data, start_row)
    new_ws.title = f"{repo_name}-{total_box_num}"
    sheets = wb._sheets

    # ✅ 核心：新 sheet 直接插到最前
    sheets.remove(new_ws)
    sheets.insert(0, new_ws)
    wb.save(output_path)