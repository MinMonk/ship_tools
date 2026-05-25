from __future__ import annotations

import logging
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage

from src.constants.dir_types import DirType
from src.utils.common_utils import root_dir


logger = logging.getLogger(__name__)

JIUFANG_DIR = Path("./ship_partner/九方")
DATA_START_ROW = 15
IMAGE_COLUMN_INDEX = 22
IMAGE_COLUMN_LETTER = "V"

GENERAL_TO_EXCEL_COLUMNS = {
    "Product_Name_En": "G",
    "Product_Name_Cn": "H",
    "Brand": "I",
    "Material": "J",
    "Usr_For": "K",
    "Model": "L",
    "HS_Code": "M",
    "Unit": "O",
    "Product_Attribute": "U",
}


def invoice_value(row, key):
    return row.get(key)


def clean_cell(value: Any) -> str:
    return "" if value is None else str(value).strip()


def maybe_number(value: Any) -> Any:
    value = clean_cell(value)
    if value == "":
        return ""
    try:
        number = float(value)
    except ValueError:
        return value
    return int(number) if number.is_integer() else number


def column_width_to_pixels(width: float | None) -> int:
    width = 8.43 if width is None else width
    return int(width * 7 + 5)


def row_height_to_pixels(height: float | None) -> int:
    height = 15 if height is None else height
    return int(height * 96 / 72)


def fit_image_to_cell(image_path: Path, max_width: int, max_height: int) -> ExcelImage:
    with PILImage.open(image_path) as source:
        original_width, original_height = source.size

    scale = min(max_width / original_width, max_height / original_height, 1)
    image = ExcelImage(str(image_path))
    image.width = max(1, int(original_width * scale))
    image.height = max(1, int(original_height * scale))
    return image


def remove_images_in_column(ws, column_index: int, row_start: int, row_end: int) -> None:
    kept_images = []
    for image in getattr(ws, "_images", []):
        marker = getattr(getattr(image, "anchor", None), "_from", None)
        if marker is None:
            kept_images.append(image)
            continue
        image_column = marker.col + 1
        image_row = marker.row + 1
        if image_column == column_index and row_start <= image_row <= row_end:
            continue
        kept_images.append(image)
    ws._images = kept_images


def find_jiufang_excel(shipment_id: str, partner_dir: Path = JIUFANG_DIR) -> list[Path]:
    if not partner_dir.exists():
        return []
    return sorted(
        path for path in partner_dir.glob(f"*{shipment_id}.xlsx")
        if path.is_file() and path.stem.endswith(shipment_id)
    )


def validate_jiufang_partner_files(invoice_data, carrier_name: str, partner_dir: Path = JIUFANG_DIR) -> dict[str, Path]:
    """
    对开票信息里的九方记录做前置校验，确保每个货件有且只有一个匹配 Excel，且 C4 匹配货件 ID。
    """
    matched_files: dict[str, Path] = {}
    warnings: list[str] = []

    for invoice_record in invoice_data:
        shipment_id = clean_cell(invoice_value(invoice_record, "货件ID"))
        invoice_type = clean_cell(invoice_value(invoice_record, "物流商"))
        if invoice_type != carrier_name:
            continue

        matches = find_jiufang_excel(shipment_id, partner_dir)
        if not matches:
            warnings.append(f"【{shipment_id}】在 {partner_dir} 下找不到以货件 ID 结尾的 .xlsx 文件")
            continue
        if len(matches) > 1:
            names = ", ".join(path.name for path in matches)
            warnings.append(f"【{shipment_id}】在 {partner_dir} 下找到多个匹配 Excel: {names}")
            continue

        excel_path = matches[0]
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            ws = wb.active
            fba_code = clean_cell(ws["C4"].value)
            wb.close()
        except Exception as exc:
            warnings.append(f"【{shipment_id}】读取九方 Excel 失败: {excel_path.name}，原因: {exc}")
            continue

        if fba_code != shipment_id:
            warnings.append(f"【{shipment_id}】九方 Excel {excel_path.name} 的 C4 为 {fba_code!r}，与货件 ID 不一致")
            continue

        matched_files[shipment_id] = excel_path

    if warnings:
        for warning in warnings:
            logger.warning(warning)
        sys.exit("九方资料前置校验失败，请修正以上问题后重新执行。")

    return matched_files


def validate_basic_fields(global_info: dict[str, Any]) -> None:
    missing = [key for key in GENERAL_TO_EXCEL_COLUMNS if not clean_cell(global_info.get(key))]
    if missing:
        raise ValueError(f"basic_data/data.txt 缺少九方必填字段: {', '.join(missing)}")


def fill_jiufang_template(global_info, product_dict, package_dict, invoice_record, excel_path: Path) -> str:
    validate_basic_fields(global_info)

    shipment_id = clean_cell(invoice_value(invoice_record, "货件ID"))
    reference_id = clean_cell(invoice_value(invoice_record, "追踪编号"))
    if not reference_id:
        raise ValueError(f"【{shipment_id}】找不到 Reference ID（开票信息.xlsx 的追踪编号为空）")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    fba_code = clean_cell(ws["C4"].value)
    if fba_code != shipment_id:
        raise ValueError(f"【{shipment_id}】九方 Excel {excel_path.name} 的 C4 为 {fba_code!r}，与货件 ID 不一致")

    last_row = ws.max_row
    remove_images_in_column(ws, IMAGE_COLUMN_INDEX, DATA_START_ROW, last_row)
    image_column_width = max(column_width_to_pixels(ws.column_dimensions[IMAGE_COLUMN_LETTER].width) - 8, 1)

    filled_rows = 0
    for row in range(DATA_START_ROW, last_row + 1):
        asin = clean_cell(ws[f"N{row}"].value)
        if not asin:
            continue

        product_data = product_dict.get(asin)
        if not product_data:
            raise ValueError(f"【{shipment_id}】九方 Excel {excel_path.name} 第 {row} 行 ASIN {asin} 未在 basic_data/产品信息.csv 中找到")

        product_type = clean_cell(product_data.get("套装类型"))
        package_data = package_dict.get(product_type)
        if not package_data:
            raise ValueError(f"【{shipment_id}】ASIN {asin} 对应套装类型 {product_type!r} 未在 basic_data/产品包装信息.csv 中找到")

        listing_price = clean_cell(product_data.get("链接售卖价格"))
        if not listing_price:
            raise ValueError(f"【{shipment_id}】ASIN {asin} 的 链接售卖价格 为空，请补充 basic_data/产品信息.csv")

        image_path = Path(clean_cell(package_data.get("图片")))
        if not image_path.exists():
            raise ValueError(f"【{shipment_id}】ASIN {asin} 的图片不存在: {image_path}")

        ws[f"D{row}"] = reference_id
        for data_key, excel_column in GENERAL_TO_EXCEL_COLUMNS.items():
            ws[f"{excel_column}{row}"] = global_info[data_key]

        ws[f"P{row}"] = maybe_number(package_data.get("单箱数量"))
        ws[f"Q{row}"] = maybe_number(package_data.get("采购单价"))
        ws[f"R{row}"] = maybe_number(package_data.get("申报单价"))
        ws[f"S{row}"] = maybe_number(listing_price)

        image_height = max(row_height_to_pixels(ws.row_dimensions[row].height) - 8, 1)
        image = fit_image_to_cell(image_path, image_column_width, image_height)
        ws.add_image(image, f"{IMAGE_COLUMN_LETTER}{row}")

        filled_rows += 1

    if filled_rows == 0:
        raise ValueError(f"【{shipment_id}】九方 Excel {excel_path.name} 没有可填充的 ASIN 行")

    output_path = Path(root_dir(DirType.Invoice)) / excel_path.name
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return str(output_path)
