import os
import logging
import sys
import csv

import openpyxl
from src.invoice.process_chidao import fill_chidao_template
from src.utils.excel_utils import insert_image_into_cell, img_attr, center_style, left_newline, copy_row_style
from src.utils.common_utils import format_current_date, root_dir
from src.constants.carriers import Carrier
from src.constants.dir_types import DirType

logger = logging.getLogger(__name__)

def invoice_output_path(carrier, repo_name, total_box_num):
    """
    定义发票的输出路径
    """
    date_str = format_current_date()
    return root_dir(DirType.Invoice) + f"/{carrier}_{repo_name}_{total_box_num}_{date_str}.xlsx"

def invoice_input_path(carrier):
    """
    定义发票的数据输入路径
    """
    return f"./basic_data/template/{DirType.Invoice}/{carrier}.xlsx"

def validate_row(row):
    """
    对单行数据进行非空校验：
      - 无论物流商（row[3]）为何值，前4列（货件ID、追踪编号、是否退税、物流商）均不能为空；
      - 当物流商为 "赤道" 时，整行所有字段都不能为空；
      - 当物流商为 "海桥" 时，除了前4列，要求送货时间（row[5]）也不能为空；
      - 当物流商为 "紐酷" 时，要求渠道类型（row[4]）和送货时间（row[5]）均不能为空。
    返回一个缺失字段的列表，如果列表为空则说明该行校验通过。
    """
    headers = ["货件ID", "追踪编号", "是否退税", "物流商", "渠道类型", "送货时间", "预约船期", "货件创建时间"]
    missing = []
    # 1. 检查前4列：货件ID、追踪编号、是否退税、物流商
    for i in range(4):
        if i >= len(row) or row[i] is None or str(row[i]).strip() == '':
            missing.append(headers[i])
    
    # mode 从物流商列获取（row[3]）
    mode = row[3] if len(row) > 3 else None
    if mode == "赤道":
        # 赤道模式要求所有列都不能为空
        for i in range(len(headers)):
            if i >= len(row) or row[i] is None or str(row[i]).strip() == '':
                if headers[i] not in missing:
                    missing.append(headers[i])
    elif mode == "海桥":
        # 海桥模式要求送货时间（row[5]）不能为空
        if len(row) < 6 or row[5] is None or str(row[5]).strip() == '':
            if headers[5] not in missing:
                missing.append(headers[5])
    elif mode == "紐酷":
        # 紐酷模式要求渠道类型（row[4]）和送货时间（row[5]）均不能为空
        if len(row) < 5 or row[4] is None or str(row[4]).strip() == '':
            if headers[4] not in missing:
                missing.append(headers[4])
        if len(row) < 6 or row[5] is None or str(row[5]).strip() == '':
            if headers[5] not in missing:
                missing.append(headers[5])
    return missing

def read_excel(file_path, start_row=3):
    # 加载工作簿和默认工作表
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 从指定行开始读取数据
    valid_rows = []
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        # 检查整行是否为空（所有单元格都是None或空字符串）
        if all(cell is None or (isinstance(cell, str) and cell.strip() == "") for cell in row):
            continue  # 跳过完全空的行
        
        missing_columns = validate_row(row)
        if missing_columns:
            shipment_id = row[0] if len(row) > 0 and row[0] is not None else "未知"
            missing_str = "、".join(missing_columns)
            errors.append(f"第{idx}行货件ID为 {shipment_id} 的数据，【{missing_str}】列不能为空")
        else:
            valid_rows.append(row)

    if errors:
        error_text = "\n".join(errors)
        sys.exit(f"Excel数据校验失败，请完善信息之后再继续执行程序。存在以下错误：\n{error_text}")

    return valid_rows

def parse_shipment_csv(directory_path, product_dict):
    """
    从指定目录下查找所有 CSV 文件并逐个解析：
      1) 前 9 行为固定头信息，存入 shipment_info，格式： {第一列: 第二列, ...}
      2) 从第 10 行开始解析商品列表：
         - 第3列 (索引2) 为 ASIN；
         - 第15列 (索引14) 为箱数（转换为整数）；
         - 第16列 (索引15) 为数量（转换为整数）；
         - 通过 ASIN 在 product_dict 中获取对应的 "套装类型" 作为 product_type（不存在时设为 "Unknown"）；
         - 对相同 product_type 的数量与箱数进行累加汇总。
    返回一个字典，格式为：
       {
           shipment_info["货件编号"]: {
               "shipment_info": { ... },
               "list_data": {
                   product_type1: {"total_num": X, "total_boxes": Y},
                   product_type2: {"total_num": A, "total_boxes": B},
                   ...
               }
           },
           ...
       }
    """
    # 获取目录下所有 CSV 文件
    csv_files = [f for f in os.listdir(
        directory_path) if f.lower().endswith('.csv')]
    if not csv_files:
        raise FileNotFoundError("在指定目录下未找到任何 CSV 文件。")

    final_result = {}
    # 遍历目录中的所有 CSV 文件
    for csv_file in csv_files:
        csv_path = os.path.join(directory_path, csv_file)
        try:
            with open(csv_path, 'r', encoding='utf-8-sig') as f:
                rows = list(csv.reader(f))
        except Exception as e:
            logger.info(f"读取文件 {csv_file} 失败：{e}")
            continue

        # 如果文件行数不足 9 行，跳过该文件
        if len(rows) < 9:
            logger.info(f"文件 {csv_file} 行数不足 9 行，跳过。")
            continue

        # 解析前 9 行为 shipment_info
        shipment_info = {}
        for i in range(9):
            row = rows[i]
            if len(row) < 2:
                continue
            key = row[0].strip()
            value = row[1].strip()
            shipment_info[key] = value

        # 如果 shipment_info 为空或缺少 '货件编号'，跳过该文件
        if not shipment_info:
            logger.info(f"文件 {csv_file} 解析到的 shipment_info 为空，跳过。")
            continue
        if "货件编号" not in shipment_info:
            logger.info(f"文件 {csv_file} 未找到 '货件编号' 字段，跳过。")
            continue
        shipment_number = shipment_info["货件编号"]
        shipment_info['仓库名称'] = shipment_info["货件名称"].split('-')[-1]

        # 从第 11 行（索引10）开始解析商品列表，并按 product_type 汇总数量和箱数
        list_data = {}
        for row in rows[10:]:
            # 检查该行是否有足够的列（至少 15 列）
            if len(row) < 15:
                continue
            asin = row[2].strip()
            # 根据 ASIN 在 product_dict 中查找对应的 "套装类型"
            product_type = product_dict.get(asin, {}).get("套装类型", "Unknown")
            if product_type == "Unknown":
                raise ValueError(f"找不到{asin}的包装信息，请检查...")
            try:
                quantity = int(row[13].strip())
            except ValueError:
                quantity = 0
            try:
                boxes = int(row[12].strip())
            except ValueError:
                boxes = 0

            package_seq_num = row[14].strip().replace(shipment_number, "")

            if product_type not in list_data:
                list_data[product_type] = {
                    "total_num": 0, "total_boxes": 0, "package_seq": []}
            list_data[product_type]["total_num"] += quantity
            list_data[product_type]["total_boxes"] += boxes

            if package_seq_num:
                list_data[product_type]["package_seq"].append(package_seq_num)

        # 将解析结果存入 final_result，使用货件编号作为 key
        final_result[shipment_number] = {
            "shipment_info": shipment_info,
            "list_data": list_data
        }

    return final_result

def fill_niuku_template(global_info, package_dict, invoice_data, ship_data, start_row=21):
    # 加载模板文件
    wb = openpyxl.load_workbook(invoice_input_path(Carrier.NiuKu))
    ws = wb.active

    # 填充固定单元格
    # 客户单号
    ws["B2"] = invoice_data[0]
    # 国家
    ws["B3"] = "美国"
    # 海运/空运
    ws["B4"] = "海运"
    # 渠道名称
    ws["B5"] = invoice_data[4]
    # 是否含税
    ws["B6"] = "含税"
    # 派送类型
    ws["B7"] = "拆柜"
    # 派送方式
    ws["B8"] = "卡车"
    # 装柜方式
    ws["B9"] = "拼柜"
    # 送货仓
    ws["B11"] = "深圳总仓"
    # 送货时间
    ws["B12"] = invoice_data[5]
    # 上门提货/装柜
    ws["B13"] = "不需要"
    # 购买保价服务
    ws["B14"] = "不需要"
    # 报关方式
    ws["B15"] = "退税" if invoice_data[2] == "是" else "买单"
    # 业务类型
    ws["B16"] = "FBA地址"
    # 仓库代码
    repo_name = ship_data.get('shipment_info').get('仓库名称')
    ws["B17"] = repo_name

    # 准备图片插入相关信息
    image_errors = []
    # 为图片所在的列设置宽度（假设第3列为图片所在列，列标为 "C"）
    ws.column_dimensions['A'].width = img_attr['col_width']  # 单位大约字符数

    # 从 start_row 开始写入 ship_data['list_data'] 中的每条记录
    current_row = start_row
    total_box_num = 0;
    for product_type, data in ship_data.get('list_data', {}).items():
        # 第 1 列：图片插入
        # 根据套装类型获取装箱信息，含图片信息
        package_data = package_dict.get(product_type)
        insert_image_into_cell(ws, package_data, current_row, img_attr, image_errors, cell_anchor="A")
        # 第 2 列：中文品名  取自./basic_data/data.txt
        ws.cell(row=current_row, column=2, value=global_info.get('Product_Name_Cn'))
        # 第 3 列：英文品名 固定  取自./basic_data/data.txt
        ws.cell(row=current_row, column=3, value=global_info.get('Product_Name_En'))
        # 第 5 列：材质  取自./basic_data/data.txt
        ws.cell(row=current_row, column=5, value=global_info.get('Material'))
        # 第 6 列：用途  取自./basic_data/data.txt
        ws.cell(row=current_row, column=6, value=global_info.get('Usr_For'))
        # 第 7 列：海关编码  取自./basic_data/data.txt
        ws.cell(row=current_row, column=7, value=global_info.get('HS_Code'))
        # 第 8 列：产品分类  固定值：否
        ws.cell(row=current_row, column=8, value='否')
        # 第 9 列：品牌  取自./basic_data/data.txt
        ws.cell(row=current_row, column=9, value=global_info.get('Brand'))
        # 第 10 列：品牌类型  固定值：境内自主品牌
        ws.cell(row=current_row, column=10, value='境内自主品牌')
        # 第 11 列：型号  固定值：无
        ws.cell(row=current_row, column=11, value='无')
        # 第 12 列：总数量
        ws.cell(row=current_row, column=12, value=data.get('total_num'))
        # 第 13 列：产品数量单位  固定值：盒
        ws.cell(row=current_row, column=13, value='盒')
        # 第 14 列：清关申报单价  取自./basic_data/产品包装信息.csv
        cell = ws.cell(row=current_row, column=14, value=int(package_data.get('申报单价').strip()))
        cell.number_format = "0"
        # 第 15 列：清关申报总金额  公式 "=L{row}*N{row}"
        cell = ws.cell(row=current_row, column=15, value=f"=L{current_row}*N{current_row}")
        cell.number_format = "0"
        # 第 16 列：采购单价  取自./basic_data/产品包装信息.csv
        cell = ws.cell(row=current_row, column=16, value=int(package_data.get('采购单价').strip()))
        cell.number_format = "0"
        # 第 17 列：采购总金额  公式 "=L{row}*P{row}"
        ws.cell(row=current_row, column=17, value=f"=L{current_row}*P{current_row}")
        cell.number_format = "0"
        # 第 18 列：生产地址  固定值：深圳
        ws.cell(row=current_row, column=18, value='深圳')

        # 第 20 列：箱数（total_boxes）
        boxes = data.get('total_boxes')
        total_box_num += boxes
        cell = ws.cell(row=current_row, column=20, value=boxes)
        cell.number_format = "0"

        # 第 21-23 列：长、宽、高  取自取自./basic_data/产品包装信息.csv
        ws.cell(row=current_row, column=21, value=float(package_data.get('长(cm)')))
        ws.cell(row=current_row, column=22, value=float(package_data.get('宽(cm)')))
        ws.cell(row=current_row, column=23, value=float(package_data.get('高(cm)')))
        # 第 24 列：总体积  公式 "= 长*宽*高*箱数/1000000"
        cell = ws.cell(row=current_row, column=24, value=f"=T{current_row}*U{current_row}*V{current_row}*W{current_row}/1000000")
        cell.number_format = "0.00"
        # 第 25 列：总净重  单箱重量 * 箱数
        net_weight = package_data.get('净重(kg)')
        cell = ws.cell(row=current_row, column=25)  # 获取单元格对象
        cell.value = round(float(net_weight) * int(boxes), 1)
        cell.number_format = "0.00"
        # 第 26 列：总毛重  单箱重量 * 箱数
        gross_weight = package_data.get('毛重(kg)')
        cell = ws.cell(row=current_row, column=26)
        cell.value = round(float(gross_weight) * int(boxes), 1)
        cell.number_format = "0.00"
        # 第 27 列：发票中的货件编号
        ws.cell(row=current_row, column=27, value=invoice_data[0])
        # 第 28 列：跟踪编号  取自./开票信息.csv
        ws.cell(row=current_row, column=28, value=invoice_data[1])
        # 第 29 列：货件创建时间  取自./开票信息.csv
        ws.cell(row=current_row, column=29, value=invoice_data[7])
        # 第 30 列：FBA箱号  取自ship_data['list_data']
        package_seq = ",".join(data.get('package_seq'))
        ws.cell(row=current_row, column=30, value=package_seq)
        # 第 31 列：留空，不做处理

        current_row += 1
        # 复制前一行的样式
        copy_row_style(ws, start_row, current_row)

    # 确保输出目录存在  仓库名称_总箱数_yyyyMMdd.xlsx
    output_path = invoice_output_path("紐酷", repo_name, total_box_num)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # 保存工作簿到输出路径
    wb.save(output_path)
    logger.debug(f" 数据已保存到 {output_path}")

    # 输出图片加载错误信息（如果有）
    if image_errors:
        logger.info("以下行图片加载失败：")
        for err in image_errors:
            logger.info(err)

def fill_pengchenghai_template(global_info, package_dict, invoice_data, ship_data, start_row=9):
    # 加载模板文件
    wb = openpyxl.load_workbook(invoice_input_path(Carrier.PengChengHai))
    ws = wb.active

    # 填充固定单元格
    repo_name = ship_data.get('shipment_info').get('仓库名称')
    # 亚马逊仓库名称
    ws["B3"] = f"{repo_name} - {ship_data.get('shipment_info').get('配送地址')}\nUS"

    # 准备图片插入相关信息
    image_errors = []
    # 为图片所在的列设置宽度
    ws.column_dimensions['U'].width = img_attr['col_width']  # 单位大约字符数

    # 从 start_row 开始写入 ship_data['list_data'] 中的每条记录
    current_row = start_row
    total_box_num = 0;
    for product_type, data in ship_data.get('list_data', {}).items():
        # 第 1 列：发票中的货件编号
        ws.cell(row=current_row, column=1, value=invoice_data[0])
        # 第 2 列：跟踪编号  取自./开票信息.csv
        ws.cell(row=current_row, column=2, value=invoice_data[1])
        # 第 3 列：中文品名  取自./basic_data/data.txt
        ws.cell(row=current_row, column=3,
                value=global_info.get('Product_Name_Cn'))
        # 第 4 列：英文品名 固定  取自./basic_data/data.txt
        ws.cell(row=current_row, column=4,
                value=global_info.get('Product_Name_En'))
        # 第 5 列：材质  取自./basic_data/data.txt
        ws.cell(row=current_row, column=5, value=global_info.get('Material'))
        # 第 6 列：用途  取自./basic_data/data.txt
        ws.cell(row=current_row, column=6, value=global_info.get('Usr_For'))
        # 第 7 列：海关编码  取自./basic_data/data.txt
        ws.cell(row=current_row, column=7, value=global_info.get('HS_Code'))
        # 第 8 列：品牌  取自./basic_data/data.txt
        ws.cell(row=current_row, column=8, value=global_info.get('Brand'))
        # 第 9 列：型号  固定值： 无
        ws.cell(row=current_row, column=9, value='无')
        # 第 10 列：总数量  公式 "=M{row}*N{row}" 单箱数量 * 箱数
        ws.cell(row=current_row, column=10,
                value=f"=M{current_row}*N{current_row}")
        # 第 11 列：清关申报单价  取自./basic_data/产品包装信息.csv
        package_data = package_dict.get(product_type)
        ws.cell(row=current_row, column=11, value=package_data.get('申报单价'))
        # 第 12 列：清关申报总金额  公式 "=J{row}*K{row}" （L列为第12列，O列为第15列）
        ws.cell(row=current_row, column=12,
                value=f"=J{current_row}*K{current_row}")
        # 第 13 列：单箱数量  取自取自./basic_data/产品包装信息.csv
        ws.cell(row=current_row, column=13, value=package_data.get('单箱数量'))
        # 第 14 列：箱数（total_boxes）
        boxes = data.get('total_boxes')
        total_box_num += boxes
        ws.cell(row=current_row, column=14, value=boxes)
        # 第 15 列：总净重  单箱重量 * 箱数
        net_weight = package_data.get('净重(kg)')
        ws.cell(row=current_row, column=15, value=round(
            float(net_weight) * int(boxes), 1))
        # 第 16 列：总毛重  单箱重量 * 箱数
        gross_weight = package_data.get('毛重(kg)')
        ws.cell(row=current_row, column=16, value=round(
            float(gross_weight) * int(boxes), 1))
        # 第 17-19 列 长、宽、高  取自取自./basic_data/产品包装信息.csv
        ws.cell(row=current_row, column=17, value=package_data.get('长(cm)'))
        ws.cell(row=current_row, column=18, value=package_data.get('宽(cm)'))
        ws.cell(row=current_row, column=19, value=package_data.get('高(cm)'))
        # 第 20 列：总体积  公式 "= 长*宽*高*箱数/1000000"
        ws.cell(row=current_row, column=20,
                value=f"=Q{current_row}*R{current_row}*S{current_row}*N{current_row}/1000000")
        # 第 21 列：图片插入
        # 根据套装类型获取装箱信息，含图片信息
        package_data = package_dict.get(product_type)
        insert_image_into_cell(ws, package_data, current_row, img_attr, image_errors, cell_anchor="U")

        current_row += 1

    # 确保输出目录存在  仓库名称_总箱数_yyyyMMdd.xlsx
    output_path = invoice_output_path("鹏城海", repo_name, total_box_num)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # 保存工作簿到输出路径
    wb.save(output_path)
    logger.debug(f" 数据已保存到 {output_path}")

    # 输出图片加载错误信息（如果有）
    if image_errors:
        logger.info("以下行图片加载失败：")
        for err in image_errors:
            logger.info(err)

def fill_yigang_template(global_info, package_dict, invoice_data, ship_data, start_row=18):
    # 加载模板文件
    wb = openpyxl.load_workbook(invoice_input_path(Carrier.YiGang))
    ws = wb.active

    # 填充固定单元格
    # 客户单号
    ws["B1"] = invoice_data[0]
    # 仓库代码
    shipment_info = ship_data.get('shipment_info')
    repo_name = shipment_info.get('仓库名称')
    ws["B3"] = repo_name
    # 收件人地址
    address = shipment_info.get('配送地址')
    ws["B6"] = f"{repo_name} - {address}, US"
    # 报关方式
    ws["F6"] = "报关退税" if invoice_data[2] == "是" else "买单"

    # 准备图片插入相关信息
    image_errors = []
    # 为图片所在的列设置宽度
    ws.column_dimensions['R'].width = img_attr['col_width']  # 单位大约字符数

    # 从 start_row 开始写入 ship_data['list_data'] 中的每条记录
    current_row = start_row
    total_box_num = 0;

    for product_type, data in ship_data.get('list_data', {}).items():

        # 第 1 列：FBA箱号  取自ship_data['list_data']
        package_seq = ",".join(data.get('package_seq'))
        ws.cell(row=current_row, column=1, value=package_seq)
        cell_1 = ws.cell(row=current_row, column=1)
        cell_1.alignment = left_newline
        # 第 2 列：跟踪编号  取自./开票信息.csv
        ws.cell(row=current_row, column=2, value=invoice_data[1])
        # 第 3 列：总毛重  单箱重量 * 箱数
        boxes = data.get('total_boxes')
        total_box_num += boxes
        package_data = package_dict.get(product_type)
        gross_weight = package_data.get('毛重(kg)')
        ws.cell(row=current_row, column=3, value=round(
            float(gross_weight) * int(boxes), 1))
        # 第 4-6 列：长、宽、高  取自取自./basic_data/产品包装信息.csv
        ws.cell(row=current_row, column=4,
                value=float(package_data.get('长(cm)')))
        ws.cell(row=current_row, column=5,
                value=float(package_data.get('宽(cm)')))
        ws.cell(row=current_row, column=6,
                value=float(package_data.get('高(cm)')))
        # 第 7 列：英文品名 固定  取自./basic_data/data.txt
        ws.cell(row=current_row, column=7,
                value=global_info.get('Product_Name_En'))
        # 第 8 列：中文品名  取自./basic_data/data.txt
        ws.cell(row=current_row, column=8,
                value=global_info.get('Product_Name_Cn'))
        # 第 9 列：清关申报单价  取自./basic_data/产品包装信息.csv
        ws.cell(row=current_row, column=9, value=package_data.get('申报单价'))
        # 第 10 列：申报数量
        ws.cell(row=current_row, column=10, value=data.get('total_num'))
        # 第 11 列：材质  取自./basic_data/data.txt
        ws.cell(row=current_row, column=11, value=global_info.get('Material'))
        # 第 12 列：海关编码  取自./basic_data/data.txt
        ws.cell(row=current_row, column=12, value=global_info.get('HS_Code'))
        # 第 13 列：用途  取自./basic_data/data.txt
        ws.cell(row=current_row, column=13, value=global_info.get('Usr_For'))
        # 第 14 列：品牌  取自./basic_data/data.txt
        ws.cell(row=current_row, column=14, value=global_info.get('Brand'))
        cell_14 = ws.cell(row=current_row, column=14)
        cell_14.alignment = center_style
         # 第 15 列：产品型号  固定值：无
        ws.cell(row=current_row, column=15, value='无')
        cell_15 = ws.cell(row=current_row, column=15)
        cell_15.alignment = center_style

        # 第 18 列：图片插入
        # 根据套装类型获取装箱信息，含图片信息
        package_data = package_dict.get(product_type)
        insert_image_into_cell(ws, package_data, current_row, img_attr, image_errors, cell_anchor="R")
        # 其它列：留空，不做处理

        current_row += 1

    # 总箱数
    ws["B16"] = total_box_num

    # 确保输出目录存在  仓库名称_总箱数_yyyyMMdd.xlsx
    output_path = invoice_output_path("壹港", repo_name, total_box_num)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # 保存工作簿到输出路径
    wb.save(output_path)
    logger.debug(f" 数据已保存到 {output_path}")

    # 输出图片加载错误信息（如果有）
    if image_errors:
        logger.info("以下行图片加载失败：")
        for err in image_errors:
            logger.info(err)

def fill_bailitong_template(global_info, package_dict, invoice_data, ship_data, start_row=21):
    # 加载模板文件
    wb = openpyxl.load_workbook(invoice_input_path(Carrier.BaiLiTong))
    ws = wb.active

    # 填充固定单元格
    # 客户单号
    ws["B2"] = invoice_data[0]
    # 报关方式
    ws["B4"] = "报关退税" if invoice_data[2] == "是" else "买单报关"
    # 仓库代码
    shipment_info = ship_data.get('shipment_info')
    repo_name = shipment_info.get('仓库名称')
    ws["B9"] = repo_name
     # 收件人地址
    address = shipment_info.get('配送地址')
    ws["B12"] = f"{repo_name} - {address}, US"

    # 准备图片插入相关信息
    image_errors = []
    # 为图片所在的列设置宽度
    ws.column_dimensions['R'].width = img_attr['col_width']  # 单位大约字符数

    # 从 start_row 开始写入 ship_data['list_data'] 中的每条记录
    current_row = start_row
    total_box_num = 0;

    for product_type, data in ship_data.get('list_data', {}).items():
        # 第 1 列：FBA箱号  取自ship_data['list_data']
        package_seq = ",".join(data.get('package_seq'))
        ws.cell(row=current_row, column=1, value=package_seq)
        cell_1 = ws.cell(row=current_row, column=1)
        cell_1.alignment = left_newline
        # 第 2 列：跟踪编号  取自./开票信息.csv
        ws.cell(row=current_row, column=2, value=invoice_data[1])
        # 第 3 列：英文品名 固定  取自./basic_data/data.txt
        ws.cell(row=current_row, column=3,
                value=global_info.get('Product_Name_En'))
        # 第 4 列：中文品名  取自./basic_data/data.txt
        ws.cell(row=current_row, column=4,
                value=global_info.get('Product_Name_Cn'))
        # 第 5 列：海关编码  取自./basic_data/data.txt
        ws.cell(row=current_row, column=5, value=global_info.get('HS_Code'))
        # 第 6 列：材质Cn  取自./basic_data/data.txt
        material = global_info.get('Material').split("/")
        ws.cell(row=current_row, column=6, value=material[1])
        # 第 7 列：材质En  取自./basic_data/data.txt
        ws.cell(row=current_row, column=7, value=material[0])
        # 第 8 列：用途  取自./basic_data/data.txt
        ws.cell(row=current_row, column=8, value=global_info.get('Usr_For'))
        # 第 9 列：单箱数量  取自取自./basic_data/产品包装信息.csv
        package_data = package_dict.get(product_type)
        ws.cell(row=current_row, column=9, value=package_data.get('单箱数量'))
        # 第 10 列：清关申报单价  取自./basic_data/产品包装信息.csv
        de_price = package_data.get('申报单价')
        ws.cell(row=current_row, column=10, value=de_price)
        # 第 11 列：申报总价
        total_num = data.get('total_num')
        boxes = data.get('total_boxes')
        total_box_num += boxes
        ws.cell(row=current_row, column=11, value=round(
            float(de_price) * int(total_num), 1))
        # 第 12 列：单箱毛重
        ws.cell(row=current_row, column=11, value=package_data.get('毛重(kg)'))
        # 第 13-15 列：长、宽、高  取自取自./basic_data/产品包装信息.csv
        ws.cell(row=current_row, column=13,
                value=float(package_data.get('长(cm)')))
        ws.cell(row=current_row, column=14,
                value=float(package_data.get('宽(cm)')))
        ws.cell(row=current_row, column=15,
                value=float(package_data.get('高(cm)')))
        # 第 16 列：品牌  取自./basic_data/data.txt
        ws.cell(row=current_row, column=16, value=global_info.get('Brand'))
        cell_16 = ws.cell(row=current_row, column=16)
        cell_16.alignment = center_style
        # 第 17 列：产品型号  固定值：无
        ws.cell(row=current_row, column=17, value='无')
        # 第 18 列：图片插入
        # 根据套装类型获取装箱信息，含图片信息
        package_data = package_dict.get(product_type)
        insert_image_into_cell(ws, package_data, current_row, img_attr, image_errors, cell_anchor="R")

        current_row += 1
        # 复制前一行的样式
        copy_row_style(ws, start_row, current_row)

    # 总箱数
    ws["B19"] = total_box_num

    # 确保输出目录存在  仓库名称_总箱数_yyyyMMdd.xlsx
    output_path = invoice_output_path("佰利通", repo_name, total_box_num)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # 保存工作簿到输出路径
    wb.save(output_path)
    logger.debug(f" 数据已保存到 {output_path}")

    # 输出图片加载错误信息（如果有）
    if image_errors:
        logger.info("以下行图片加载失败：")
        for err in image_errors:
            logger.info(err)

def fill_yitonganda_template(global_info, package_dict, invoice_data, ship_data, start_row=18):
    """易通安达"""
    # 加载模板文件
    wb = openpyxl.load_workbook(invoice_input_path(Carrier.YiTongAnDa))
    ws = wb.active

    # 填充固定单元格
    # 客户单号
    ws["B1"] = invoice_data[0]
    ws["D1"] = invoice_data[0]
    # 交货仓库
    ws["B4"] = "深圳龙华仓"
    # 走货路线
    ws["B4"] = invoice_data[4]
    # 报关方式
    ws["B5"] = "单独报关" if invoice_data[2] == "是" else "买单报关"
    # 仓库代码
    shipment_info = ship_data.get('shipment_info')
    repo_name = shipment_info.get('仓库名称')
    ws["B6"] = repo_name
    # 地址类型
    ws["B7"] = "亚马逊地址"
    ws["B8"] = repo_name


    # 准备图片插入相关信息
    image_errors = []
    # 为图片所在的列设置宽度
    ws.column_dimensions['N'].width = img_attr['col_width']  # 单位大约字符数

    # 从 start_row 开始写入 ship_data['list_data'] 中的每条记录
    current_row = start_row
    total_box_num = 0
    total_volume = 0

    for product_type, data in ship_data.get('list_data', {}).items():

        # 根据套装类型获取装箱信息，含图片信息
        package_data = package_dict.get(product_type)

        # 第 1 列：FBA箱号  取自ship_data['list_data']
        package_seq = ",".join(data.get('package_seq'))
        ws.cell(row=current_row, column=1, value=package_seq)
        cell_1 = ws.cell(row=current_row, column=1)
        cell_1.alignment = left_newline
        # 第 2 列：跟踪编号  取自./开票信息.csv
        ws.cell(row=current_row, column=2, value=invoice_data[1])
        # 第 3 列：中文品名  取自./basic_data/data.txt
        ws.cell(row=current_row, column=3, value=global_info.get('Product_Name_Cn'))
        # 第 4 列：英文品名 固定  取自./basic_data/data.txt
        ws.cell(row=current_row, column=4, value=global_info.get('Product_Name_En'))
        # 第 5 列：材质  取自./basic_data/data.txt
        ws.cell(row=current_row, column=5, value=global_info.get('Material'))
        # 第 6 列：用途  取自./basic_data/data.txt
        ws.cell(row=current_row, column=6, value=global_info.get('Usr_For'))
        # 第 7 列：海关编码  取自./basic_data/data.txt
        ws.cell(row=current_row, column=7, value=global_info.get('HS_Code'))
        # 第 8 列：产品类型  固定值：普货
        ws.cell(row=current_row, column=8, value='普货')
        # 第 9 列：单箱产品数量
        ws.cell(row=current_row, column=9, value=package_data.get('单箱数量'))
        # 第 10 列：箱数
        boxes = data.get('total_boxes')
        total_box_num += boxes
        ws.cell(row=current_row, column=10, value=boxes)
        # 第 11 列：币种  固定值：美金
        ws.cell(row=current_row, column=11, value='美金')
        # 第 12 列：清关申报单价  取自./basic_data/产品包装信息.csv
        cell = ws.cell(row=current_row, column=12, value=float(package_data.get('申报单价').strip()))
        cell.number_format = "0"
        # 第 13 列：单个产品净重
        ws.cell(row=current_row, column=13, value=float(package_data.get('单产品重量(kg)').strip()))
        # 第 14 列：图片插入
        insert_image_into_cell(ws, package_data, current_row, img_attr, image_errors, cell_anchor="N")
        # 第 15 列：品牌  取自./basic_data/data.txt
        ws.cell(row=current_row, column=15, value=global_info.get('Brand'))
        cell_15 = ws.cell(row=current_row, column=15)
        cell_15.alignment = center_style
        # 第 16 列：品牌类型  固定值：境内自主品牌
        ws.cell(row=current_row, column=16, value='境内自主品牌')
        # 第 17 列：型号  固定值：无
        ws.cell(row=current_row, column=17, value='无')
        cell_17 = ws.cell(row=current_row, column=17)
        cell_17.alignment = center_style
        # 第 20 列：总产品数量
        ws.cell(row=current_row, column=20, value=f"=I{current_row}*J{current_row}")
        # 第 21 列：申报总金额  公式 "=L{row}*T{row}" 
        ws.cell(row=current_row, column=21, value=f"=L{current_row}*T{current_row}")
        # 第 23 列：实重
        ws.cell(row=current_row, column=23, value=float(package_data.get('毛重(kg)')))
        # 第 24-26 列：长、宽、高  取自取自./basic_data/产品包装信息.csv
        length = float(package_data.get('长(cm)'))
        width = float(package_data.get('宽(cm)'))
        height = float(package_data.get('高(cm)'))
        ws.cell(row=current_row, column=24, value=length)
        ws.cell(row=current_row, column=25, value=width)
        ws.cell(row=current_row, column=26, value=height)

        # 计算体积
        single_single = length * width * height * boxes / 1000000
        total_volume += single_single
        # 其它列：留空，不做处理

        current_row += 1
        # 复制前一行的样式
        copy_row_style(ws, start_row, current_row)

    # 总件数
    ws.cell(row=current_row, column=20, value=f"=SUM(T{start_row}:T{current_row-1})")
    # 总申报额
    ws.cell(row=current_row, column=21, value=f"=SUM(U{start_row}:U{current_row-1})")
    # 总箱数
    ws["B16"] = total_box_num
    # 总体积
    ws["D15"] = total_volume

    # 确保输出目录存在  仓库名称_总箱数_yyyyMMdd.xlsx
    output_path = invoice_output_path("易通安达", repo_name, total_box_num)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # 保存工作簿到输出路径
    wb.save(output_path)
    logger.debug(f" 数据已保存到 {output_path}")

    # 输出图片加载错误信息（如果有）
    if image_errors:
        logger.info("以下行图片加载失败：")
        for err in image_errors:
            logger.info(err)

def run_invoice(global_info, product_dict, package_dict):
    """
    Invoice 功能入口：解析 & 校验 开票信息 → 解析装箱数据 → 调用各渠道模板填充
    """

    # 加载并验证 invoice 信息
    invoice_data = read_excel('./plan_data/开票信息.xlsx')
    logger.info('【开票信息】解析 & 校验完成...')

    # 加载 Amazon 后台下载的装箱数据
    ship_data = parse_shipment_csv('./shipment_data/', product_dict)
    logger.info('【装箱数据】加载完成...')

    # 基于./开票信息.csv 文件中维护的数据来生成发票
    for invoice_record in invoice_data:
        shipment_id = invoice_record[0]
        invoice_type = invoice_record[3]
        repo_name = invoice_record[8]
        if invoice_type == Carrier.NiuKu:
            fill_niuku_template(global_info, package_dict, invoice_record, ship_data.get(shipment_id))
            logger.info(f"【{shipment_id}】-【{repo_name}】开具[紐酷]发票成功...")
        elif invoice_type == Carrier.PengChengHai:
            fill_pengchenghai_template(global_info, package_dict, invoice_record, ship_data.get(shipment_id))
            logger.info(f"【{shipment_id}】-【{repo_name}】开具[鹏城海]发票成功...")
        elif invoice_type == Carrier.YiGang:
            fill_yigang_template(global_info, package_dict, invoice_record, ship_data.get(shipment_id))
            logger.info(f"【{shipment_id}】-【{repo_name}】开具[壹港]发票成功...")
        elif invoice_type == Carrier.BaiLiTong:
            fill_bailitong_template(global_info, package_dict, invoice_record, ship_data.get(shipment_id))
            logger.info(f"【{shipment_id}】-【{repo_name}】开具[百利通]发票成功...")
        elif invoice_type == Carrier.ChiDao:
            fill_chidao_template(global_info, package_dict, invoice_record, ship_data.get(shipment_id))
            logger.info(f"【{shipment_id}】-【{repo_name}】开具[赤道]发票成功...")
        elif invoice_type == Carrier.YiTongAnDa:
            fill_yitonganda_template(global_info, package_dict, invoice_record, ship_data.get(shipment_id))
            logger.info(f"【{shipment_id}】-【{repo_name}】开具[易通安达]发票成功...")
        else:
            logger.info(f"未知的发票类型: {invoice_type}")
