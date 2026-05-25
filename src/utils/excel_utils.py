import os
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from copy import copy


# 样式
center_style = Alignment(horizontal='center', vertical='center', wrap_text=True)  # 居中对齐
left_newline = Alignment(horizontal='left', vertical='center', wrap_text=True)  # 左对齐，自动换行
img_attr = {'col_width': 22, 'row_height': 102, 'img_width': 145, 'img_height': 133}

def insert_image_into_cell(ws, package_data, row, img_attr, errors, cell_anchor="U"):
    """
    在指定单元格插入图片；如果失败，记录行号
    """
    pic_path = package_data.get('图片')
    if not pic_path or not os.path.exists(pic_path):
        errors.append(f'Row {row}: image not found for {package_data}, {pic_path}')
        return
    img = Image(pic_path)
    img.width = img_attr['img_width']
    img.height = img_attr['img_height']
    anchor = f"{cell_anchor}{row}"
    ws.add_image(img, anchor)
    # 调整行高和列宽
    ws.row_dimensions[row].height = img_attr['row_height']
    ws.column_dimensions[cell_anchor].width = img_attr['col_width']


def clean_sheet_data(ws, start_row: int, end_row: int, exclude_col):
    """
    通用清空方法：
    - 清空 ws 在 [start_row, end_row] 行区间内的数据
    - exclude_col: 数组，支持 ["A","B","AA"] 这种列名，也支持 [1,2,27] 这种列索引（1-based），也可混用
    - 默认清空范围：从第 1 列到 ws.max_column（你也可以根据需要改成固定列数）
    """
    if start_row > end_row:
        return

    # 兼容 exclude_col 为 None/空
    exclude_col = exclude_col or []

    def col_letter_to_index(col_letters: str) -> int:
        """Excel 列名(如 A/AB) -> 1-based index"""
        col_letters = col_letters.strip().upper()
        if not col_letters.isalpha():
            raise ValueError(f"Invalid column letters: {col_letters}")
        idx = 0
        for ch in col_letters:
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
        return idx

    exclude_set = set()
    for c in exclude_col:
        if isinstance(c, int):
            if c <= 0:
                raise ValueError(f"Invalid column index (must be 1-based): {c}")
            exclude_set.add(c)
        elif isinstance(c, str):
            exclude_set.add(col_letter_to_index(c))
        else:
            raise TypeError(f"exclude_col supports int or str, got: {type(c)}")

    # 清图片（可选但建议：复制 sheet 会带旧图，清空数据时通常也应清）
    try:
        ws._images = []
    except Exception:
        pass

    max_col = ws.max_column  # 默认清到当前 sheet 的最大列
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            if c in exclude_set:
                continue
            ws.cell(row=r, column=c).value = None

def copy_row_style(ws, source_row, target_row):
    """
    复制整行样式（高度、列宽、单元格样式）
    """
    # 行高
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    # 样式
    for col in ws.iter_cols(min_row=source_row, max_row=source_row):
        for cell in col:
            new_cell = ws.cell(row=target_row, column=cell.column)
            if cell.has_style:
                new_cell._style = copy(cell._style)
    # 列宽也保持一致
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = \
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width