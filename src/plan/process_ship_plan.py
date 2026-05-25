import logging
import os
import math
from typing import List, Tuple, Dict, Any

import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side
from src.utils.common_utils import parse_csv_to_dict2, format_current_date, root_dir
from src.constants.dir_types import DirType


logger = logging.getLogger(__name__)

# 单位转换系数
CM_TO_IN = 0.393700787  # 厘米转英寸
KG_TO_LB = 2.20462262   # 千克转磅

# ERP 模板常量
ERP_SHEET = 'Worksheet'
ERP_TEMPLATE_PATH = './basic_data/template/plan/LingXing_ERP_Shipment_Plan.xlsx'

def plan_output_path(country=None):
    """
    定义发票的输出路径
    """
    date_str = format_current_date()
    if country is None:
        return root_dir(DirType.Plan) + f"/ERP_Shipment_Plan_{date_str}.xlsx"
    return root_dir(DirType.Plan) + f"/Amazon_Shipment_{country}_{date_str}.xlsx"


def read_plan_data(
    file_path: str = './plan_data/出货计划.xlsx'
) -> Tuple[List[Tuple[str, str, int]], str, str]:
    """
    读取出货计划 Excel 的首个 sheet，格式化输出记录、国家代码和 sheet 名称

    :param file_path: 出货计划文件路径
    :return: (records, country, sheet_name)
        records: 列表，每条为 (color, size, quantity)
        country: sheet 名最后一节（以 '-' 分割），若无则 'US'
        sheet_name: 使用的 sheet 名称
    :raises FileNotFoundError: 文件不存在
    :raises ValueError: 缺少必需列
    """
    if not os.path.exists(file_path):
        logger.error(f"文件未找到: {file_path}")
        raise FileNotFoundError(f"找不到文件: {file_path}")

    xls = pd.ExcelFile(file_path)
    sheet_name = xls.sheet_names[0]
    parts = sheet_name.split('-')
    country = parts[-1] if len(parts) > 1 and parts[-1] else 'US'

    df = xls.parse(sheet_name)
    if '颜色' not in df.columns:
        logger.error("缺少必需列: '颜色'")
        raise ValueError("缺少必需列: '颜色'")

    records: List[Tuple[str, str, int]] = []
    for _, row in df.iterrows():
        color = str(row['颜色']).strip()
        for size in df.columns[1:]:
            raw = row[size]
            if pd.notna(raw) and raw != 0:
                quantity = int(raw)
                records.append((color, str(size).strip(), quantity))
    return records, country, sheet_name


def write_amazon_plan(
    records: List[Tuple[str, str, int]],
    package_dict: Dict[str, Dict],
    country: str
) -> str:
    """
    将计划记录写入 Amazon 模板文件，从第9行开始，根据国家转换单位，并为每行添加边框

    :param records: 列表，每条为 (sku, size, quantity)
    :param package_dict: 包装信息字典
    :param country: 国家代码，如 'US'
    :param template_path: 模板文件路径
    :return: 输出文件路径
    """
    if country == "CA":
        template_path = './basic_data/template/plan/Amazon_Shipment_CA.xlsx'
    else:
        template_path = './basic_data/template/plan/Amazon_Shipment_US.xlsx'


    if not os.path.exists(template_path):
        logger.error(f"模板未找到: {template_path}")
        raise FileNotFoundError(f"模板未找到: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb['Create workflow – template']
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    start_row = 7
    for idx, (sku, size, qty) in enumerate(records):
        row = start_row + idx
        pkg = package_dict.get(size) or {}
        length_cm = float(pkg.get('长(cm)', 0) or 0)
        width_cm = float(pkg.get('宽(cm)', 0) or 0)
        height_cm = float(pkg.get('高(cm)', 0) or 0)
        weight_kg = float(pkg.get('毛重(kg)', 0) or 0)
        box_size = int(pkg.get('单箱数量', 0) or 0)

        if country.upper() == 'US':
            length = round(length_cm * CM_TO_IN, 2)
            width = round(width_cm * CM_TO_IN, 2)
            height = round(height_cm * CM_TO_IN, 2)
            weight = round(weight_kg * KG_TO_LB, 2)
        else:
            length, width, height, weight = length_cm, width_cm, height_cm, weight_kg

        # box_num = qty / box_size if box_size > 0 else 0
        # 每箱数量及校验
        if box_size > 0:
            box_num = qty / box_size
            if qty % box_size != 0:
                logger.warning(f"请确认SKU: {sku} 发货数量{qty}，单箱数量为{box_size}，该 SKU 箱数不为整数，当前已忽略该SKU")
                continue
        else:
            logger.warning(f"请确认SKU: {sku} 单箱数量为0, 当前已忽略该SKU")
            continue

        cols = [1, 2, 5, 6, 7, 8, 9, 10]
        values = [sku, qty, box_size, box_num, length, width, height, weight]
        for col, val in zip(cols, values):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border

        # 整行添加边框
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col_idx).border = border

    out_path = plan_output_path(country)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    logger.info(f"已将计划写入模板并保存到: {out_path}")
    return out_path


def write_erp_plan(
    erp_records: List[Tuple[str, str, str, int, str]],
    package_dict: Dict[str, Dict],
    sheet_name: str
) -> str:
    """
    将 ERP 发货计划写入 LingXing ERP 模板，从第2行开始，每行每个单元格添加边框

    :param erp_records: 列表，每条为 (sku, fnsku, size, quantity, color)
    :param package_dict: 包装信息字典
    :param sheet_name: 出货计划 Excel sheet 名称
    :return: 输出文件路径
    """
    if not os.path.exists(ERP_TEMPLATE_PATH):
        logger.error(f"ERP 模板未找到: {ERP_TEMPLATE_PATH}")
        raise FileNotFoundError(f"ERP 模板未找到: {ERP_TEMPLATE_PATH}")

    wb = openpyxl.load_workbook(ERP_TEMPLATE_PATH)
    ws = wb[ERP_SHEET]
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    start_row = 2
    for idx, (sku, fnsku, size, qty, color) in enumerate(erp_records):
        row = start_row + idx
        pkg = package_dict.get(size) or {}
        box_size = int(pkg.get('单箱数量', 0) or 0)
        per_box = qty / box_size if box_size > 0 else 0

        # 固定列填充及边框
        ws.cell(row=row, column=1, value=sheet_name).border = border
        ws.cell(row=row, column=2, value="原厂包装").border = border
        ws.cell(row=row, column=3, value=sku).border = border
        ws.cell(row=row, column=4, value=fnsku).border = border
        ws.cell(row=row, column=6, value=box_size).border = border
        ws.cell(row=row, column=7, value=per_box).border = border
        ws.cell(row=row, column=8, value=qty).border = border
        ws.cell(row=row, column=9, value="海派").border = border
        ws.cell(row=row, column=12, value="否").border = border
        ws.cell(row=row, column=13, value="SEMAXE深圳").border = border
        ws.cell(row=row, column=18, value=f"{size}-{color}").border = border

        # 整行添加边框
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col_idx).border = border

    out_path = plan_output_path()
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    logger.info(f"已将 ERP 计划写入模板并保存到: {out_path}")
    return out_path


def summarize_plan(
    amazon_records: List[Tuple[str, str, int]],
    package_dict: Dict[str, Dict]
) -> Dict[str, Any]:
    """
    对发货计划进行汇总：SKU 数、总数量、总箱数、总体积(m³)、总重量(kg)
    """
    sku_set = set()
    total_qty = 0
    total_boxes = 0
    total_volume_m3 = 0.0
    total_weight_kg = 0.0

    for sku, size, qty in amazon_records:
        sku_set.add(sku)
        total_qty += qty

        pkg = package_dict.get(size, {}) or {}

        box_size = int(pkg.get('单箱数量', 0) or 0)
        if box_size <= 0:
            # 单箱数量缺失/为0：无法计算箱数、体积、重量
            continue

        boxes = math.ceil(qty / box_size)
        total_boxes += boxes

        # 体积：cm³ -> m³
        length_cm = float(pkg.get('长(cm)', 0) or 0)
        width_cm  = float(pkg.get('宽(cm)', 0) or 0)
        height_cm = float(pkg.get('高(cm)', 0) or 0)
        volume_per_box_m3 = (length_cm * width_cm * height_cm) / 1_000_000
        total_volume_m3 += volume_per_box_m3 * boxes

        # 重量：按“毛重(kg)”*箱数
        weight_kg = float(pkg.get('毛重(kg)', 0) or 0)
        total_weight_kg += weight_kg * boxes

    summary = {
        '总sku数': len(sku_set),
        '总盒数': total_qty,
        '总箱数': total_boxes,
        '总体积(m³)': round(total_volume_m3, 4),
        '总重量(kg)': round(total_weight_kg, 2),
    }
    logger.info(f"计划汇总: {summary}")
    return summary

def run_plan(
    global_info: pd.DataFrame,
    package_dict: Dict[str, Dict]
) -> Tuple[str, str, Dict[str, Any]]:
    """
    plan 功能入口：
      1. 读取出货计划数据
      2. 查询产品信息 (颜色-套装类型 → SKU, FNSKU)
      3. 写入 Amazon 模板与 ERP 模板
      4. 汇总并返回结果

    :param global_info: 全局配置信息（未用）
    :param package_dict: 包装信息
    :return: (amazon_path, erp_path, summary)
    """
    records, country, sheet_name = read_plan_data()
    product_data = parse_csv_to_dict2(
        './basic_data/产品信息.csv', ['颜色', '套装类型']
    )

    amazon_records: List[Tuple[str, str, int]] = []
    erp_records: List[Tuple[str, str, str, int, str]] = []
    for color, size, qty in records:
        key = f"{color}-{size}"
        info = product_data.get(key)
        if not info:
            logger.warning(f"未找到产品信息: {key}")
            continue
        sku = info['SKU']
        fnsku = info['FNSKU']
        amazon_records.append((sku, size, qty))
        erp_records.append((sku, fnsku, size, qty, color))

    amazon_path = write_amazon_plan(amazon_records, package_dict, country)
    erp_path = write_erp_plan(erp_records, package_dict, sheet_name)
    summary = summarize_plan(amazon_records, package_dict)
    return amazon_path, erp_path, summary
